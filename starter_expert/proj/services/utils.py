import csv 
import time 
import requests 
import re 
import xlsxwriter 
import io 
import os
import json
import base64
import itertools

from . import models

from pymorphy3 import MorphAnalyzer
from openpyxl import Workbook, load_workbook
from pathlib import Path

from django.http import Http404


CURRENT_DIR = Path(__file__).resolve().parent


# приводит каждое слово в переданном тексте к начальной форме
def normalize_text(text: str) -> str:
    morph = MorphAnalyzer()
    words = re.sub(r'[^\w\s]', '', text).split(' ')
    normalized_words = []
    for word in words:
        normalized_word = morph.parse(word)[0].normal_form
        normalized_words.append(normalized_word)

    return ' '.join(normalized_words)


# записывает все данные по каждому отчету в IndexerReportData
def createReportData(report):
    nmid = report.nmid

    indexer = Indexer(nmid)
    for query in indexer.iterate_queries():
        models.IndexerReportData.objects.create(
            priority_cat=query.get('top_category'),
            keywords=query.get('keywords'),
            frequency=query.get('frequency'),
            req_depth=query.get('req_depth'),
            existence=query.get('existence'),
            place=query.get('place'),
            spot_req_depth=query.get('spot_req_depth'),
            ad_place=query.get('ad_place'),
            report=report
        )

    report.ready = True
    report.save()


# класс для работы с файлами
# открытие, чтение, запись, вся хуйня
class FileOperator:

    # для работы с csv файлами используется встроенная библиотека csv
    # для работы с excel файлами требуется установка openpyexcel
    # zip архивы пока под вопросом, возможно будут описаны чуть позже


    # метод для получения nmid товаров из эксель файла
    def iterate_nmids(self, filepath: str):
        book = load_workbook(filepath)
        sheet = book.active
        nmids = set()
        for cell in sheet.iter_rows(values_only=True):
            nmid = cell[0]
            nmids.add(nmid)

        return nmids

    # применение морфологического анализатора для csv файлика
    # содержащего топ запросы за х времени
    def rewrite_top_requests(self):
        reader = csv.reader('requests.csv')
        with open('normalized_requests.csv', 'w', encoding='cp1251', newline='') as file:
            writer = csv.writer(file)

            for query in reader:
                keywords = query[0]
                frequency = query[1]
                keywords_normalized = normalize_text(keywords)
                normalized_query = [
                    keywords_normalized, frequency
                ]

                writer.writerow(normalized_query)

    # создает файл с отчетом по определенному товару
    # выгружает данные в xlsx формате
    def create_report_buffer(self, report_id):

        # инициализируем все нужные инструменты
        buffer = io.BytesIO()
        book = xlsxwriter.Workbook(buffer)
        sheet = book.add_worksheet()

        # создаем строку с названием всех колонок
        columns = [
            'priority_cat', 'keywords', 'frequency', 'req_depth', 'existence', 'place', 'spot_req_depth', 'ad_spots',
            'ad_place'
        ]

        for i in range(len(columns)):
            sheet.write(0, i, columns[i])

        # собираем данные из бд, итерируем, заносим в лист
        data = models.IndexerReportsData.objects.filter(report_id=report_id)
        row_counter = 4
        for query in data:
            sheet.write(row_counter, 0, query.priority_cat)
            sheet.write(row_counter, 1, query.keywords)
            sheet.write(row_counter, 2, query.frequency)
            sheet.write(row_counter, 3, query.req_depth)
            sheet.write(row_counter, 4, query.existence)
            sheet.write(row_counter, 5, query.place)
            sheet.write(row_counter, 6, query.spot_req_depth)
            sheet.write(row_counter, 7, query.ad_spots)
            sheet.write(row_counter, 8, query.ad_place)
            row_counter += 1

        book.close()
        buffer.seek(0)
        return buffer

    def create_seo_collector_buffer(self, query_obj):
        """Создает файл с отчетом собранных данных с wildberries"""
        buffer = io.BytesIO()
        book = xlsxwriter.Workbook(buffer)
        sheet = book.add_worksheet()

        center = book.add_format({'align': 'center'})

        sheet.set_column(0, 0, 50)
        sheet.set_column(1, 1, 25)
        sheet.set_column(2, 2, 25)
        sheet.set_column(3, 3, 35)

        # создаем строку с названием всех колонок
        columns = [
            "Фразы", "Частотность", "Глубина", "Приоритетная категория" 
        ]

        if query_obj is None:
            raise Exception(f"SeoReport с query='{query_obj.query}' и depth='{query_obj.depth}' не существует")

        sheet.write(1, 1, "Запрос: ")
        sheet.write(1, 2, query_obj.query)
        sheet.write(2, 1, "Глубина: ")
        sheet.write(2, 2, query_obj.depth)
        row_counter = 4

        for query in query_obj.keywords.all():
            sheet.write(row_counter, 0, query.keywords)
            sheet.write(row_counter, 1, query.frequency)
            sheet.write(row_counter, 2, query.req_depth)
            sheet.write(row_counter, 3, query.top_category)
            row_counter += 1

        book.close()
        buffer.seek(0)
        return buffer

    def getRequestsData(self):

        # получаем ответ от вб с закодированными данными по миллиону топ запросов
        # декодируем текст в человеческий
        resp = requests.get('https://trending-searches.wb.ru/file?period=month').json()
        enc_data = resp['data']['file']
        data = base64.b64decode(enc_data).decode('utf-8')
        queriesAsStrs = data.split('\n')

        # проходимся по каждой строчке, предварительно сплитили по переносу
        # роспаковываем на запрос и частоту
        # создаем и добавляем словарь с распакованными данными и доп полем normalized_keywords (потребуется для другой ф-и)
        quieriesAsDicts = []
        for query in queriesAsStrs:
            try:
                # некоторые запросы содержат неразделительные запятые, например
                # "постельное 1,5 бязь комплект",40
                # делаем проверочку, далее создаем, добавляем словарь
                if query.count(',') > 1:
                    pre_keywords, frequency = query.split('",')
                    keywords = pre_keywords.strip('"')
                else:
                    keywords, frequency = query.split(',')

                quieriesAsDicts.append(
                    {
                        "keywords": keywords.replace('﻿', ''),
                        "frequency": frequency,
                        "normalized_keywords": ""
                    }
                )

            except:
                continue

        # словарь в json дамп и в файл
        dump = json.dumps(quieriesAsDicts, indent=4, ensure_ascii=False)
        with open('requests.json', 'w', encoding='utf-8') as jsonFile:
            jsonFile.write(dump)



# класс предназначен для работы с ссылками
# содержит набор статичных ссылок и шаблонов для использования api вайлдберрис
class URLOperator:

    query_categories_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&filters=xsubject&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=filters&spp=0'
    subject_base_url = 'https://static-basket-01.wb.ru/vol0/data/subject-base.json'
    card_url_template = 'https://basket-replace_me.wb.ru/vol{}/part{}/{}/info/ru/card.json'
    ad_url_template = 'https://catalog-ads.wildberries.ru/api/v5/search?keyword={}'
    filtered_by_brand_id_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&fbrand={brand_id}&page=replace_me&query={query}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'
    any_query_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&page=replace_me&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'
    any_query_req_depth_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&filters=xsubject&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=filters&spp=0'
    nmid_detail_url_template = 'https://card.wb.ru/cards/detail?appType=1&curr=rub&dest=123586150&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&spp=0&nm={}'


    # метод формирует ссылку для сбора данных с карточки определенного товара
    # для обращения к api используется id товара
    def create_card_url(self, product_id: int) -> str:
        nmid = str(product_id)
        lenght = len(nmid)
        vol = lenght - 5
        part = lenght - 3
        pre_url = self.card_url_template.format(nmid[:vol], nmid[:part], nmid)

        counter = 1


        # к сожалению, так и не разобрался, от чего зависит номер этой ебучей корзины
        # зато выяснил, что корзины нумеруются макисмум до 12 и подобрать нужное число можно перебором
        # проходимся циклом, пока апи не вернет 200
        while True:
            if counter < 10:
                num = f'0{counter}'
                url = pre_url.replace('replace_me', num)
            else:
                url = pre_url.replace('replace_me', str(counter))
            try:
                resp = requests.get(url)
                if resp.status_code == 200:
                    return url
                counter += 1
            except:
                print(f"[INFO] we'll be right back in 10 seconds. {url}")
                time.sleep(10)

    # создание ссылки для сбора информации о рекламе
    def create_ad_url(self, query: str) -> str:
        return self.ad_url_template.format(
            query.strip().replace(' ', '%20')
        )


    # это возможно чуть позже будет переделываться или вообще удалено будет
    # здесь создается ссылка для сбора данных по продавцу
    # ссылка используется для определения существования товара по определенному запросу
    def create_filtered_by_brand_url(self, query, brand_id):
        return self.filtered_by_brand_id_url_template.format(
            brand_id=brand_id,
            query=query.strip().replace(' ', '%20')
        ).replace('replace_me', '1')


    # создание ссылки для получения глубины выдачи
    def create_query_req_depth_url(self, query):
        return self.any_query_req_depth_url_template.format(
            query.strip().replace(' ', '%20')
        )


    # создает ссылку для получения информации по опреденному запросу
    # без каких либо фильтров или орграничений
    def create_query_url_template(self, query):
        return self.any_query_url_template.format(
            query.strip().replace(' ', '%20')
        )

    # создает ссылку для получения инфы от апи по категориям по определенному запросу в выдаче
    # вызывается в том случае, если топ категория по определенному запросу не была получена
    def create_query_categories_url(self, query):
        return self.query_categories_url_template.format(
            query.strip().replace(' ', '%20')
        )

    def create_nmid_detail_url(self, nmid):
        return self.nmid_detail_url_template.format(nmid)


# класс DataCollector отправляет запросы на ссылки, которые были сгенерированы с помощью класса URLOperator
# получает и парсит данные от апи, а также обрабатывает возможные ошибки
class DataCollector:

    # выдергивает описание товара из пришедшей линки
    # линку предварительно сгенерировал юрл оператор в вызове методе create_card_url
    def get_card_info(self, url):
        resp = requests.get(url).json()
        info = ''
        try:
            desc = resp.get('description', '')
            info += desc + ' '
            grouped_options_info = resp.get('grouped_options', {})
            options_info = resp.get('options', {})
            grouped_options = grouped_options_info[0]['options']
            for option in grouped_options:
                info += option['value'] + ' '

            for key in options_info:
                info += option['value'] + ' '

            return info

        except Exception as e:
            print(e, url)
            return ''

    # сбор id бренда
    def get_brand_id(self, url):
        try:
            resp = requests.get(url).json()
            return resp['data']['products'][0]['brandId']

        except Exception as e:
            print(f'some error at {url} {e}')
            return 0


    # получение глубины запроса
    def get_req_depth(self, url):
        try:
            resp = requests.get(url).json()
            return resp['data']['total']
        except:
            print(f'error at {url}')
            return 0

    # получение id товаров по определенному брэнду
    # возможно тоже будет переделано или нахуй удалено
    # nmid товара при определенном запросе нужны для проверки наличия по запросу
    def get_query_by_brand(self, url: str) -> list[str]:
        counter = 1
        ids = set()
        url.replace('replace_me', str(counter))
        while True:
            url = url.replace(f'page={counter - 1}', f'page={counter}')
            try:
                resp = requests.get(url).json()
                if 'data' in resp:
                    products = resp['data']['products']
                else:
                    products = []


                if not products and counter == 1:
                    return []

                elif not products and counter > 1:
                    return ids

                else:
                    for product in products:

                        ids.add(str(product['id']))

                    counter += 1

            except Exception as e:
                print(e)
                return []

    # возвращает все nmid товаров с первых 10 (если есть) страниц
    def get_query(self, url):
        counter = 1
        ids = []
        url = url.replace('replace_me', str(counter))

        while counter <= 10:
            url = url.replace(f'page={counter - 1}', f'page={counter}')
            try:
                resp = requests.get(url).json()
                products = resp['data']['products']
                if not products:
                    return ids

                for product in products:
                    ids.append(str(product['id']))

                counter += 1
            except Exception as e:
                time.sleep(20)
                print(e)

        return ids

    # возвращает все рекламные nmid товаров
    def get_ad(self, url):
        ad_ids = []
        try:
            resp = requests.get(url).json()
            if not resp['adverts']:
                return []

            for product in resp['adverts']:
                ad_ids.append(str(product['id']))

            return ad_ids

        except:
            print(url)
            return []

    # возвращает топ категорию из апи ответа по рекламе
    def get_top_category(self, url):
        try:
            resp = requests.get(url).json()
            if resp['prioritySubjects']:
                return resp['prioritySubjects'][0]
            else:
                return 'pass'
        except Exception as e:
            print(e, url)
            return 'Some error'

    # возвращает все категории, которые есть на вб
    def get_subject_base(self, url):
        categories = {}
        try:
            resp = requests.get(url).json()
            for parent_category in resp:
                for child_category in parent_category['childs']:
                    categories.update(
                        {
                            child_category['id']: child_category['name']
                        }
                    )

            return categories

        except Exception as e:
            return {}


    # если нет топ категории по рекламе(запрос уебанский), то этот метод возвращает категорию
    # у которой наибольшее число товаров по опреденному запросу
    def get_query_top_category(self, url):
        category = 'No category'
        count = 0
        try:
            resp = requests.get(url).json()
            for key in resp['data']['filters']:
                for item in key['items']:
                    if item['count'] > count:
                        category = item['name']
                        count = item['count']
            return category

        except Exception as e:
            print(url)
            return 'Error during category getting'


    # возвращает бренд и имя товара
    # для более подробного описания текстовой части карточки
    def get_brand_and_name(self, url):
        try:
            resp = requests.get(url).json()
            product = resp['data']['products'][0]
            name = product.get('name')
            brand = product.get('brand')
            return ' '.join([name, brand])

        except Exception as e:
            print(e, url)
            return ''


# класс проверяет/сравнивает информацию
class Checker:

    # при инициализации получаем nmid и описание
    # удаляем все знаки препинания из описания товара
    def __init__(self, nmid, desc):
        self.nmid = str(nmid)
        self.desc = re.sub(r'[^\w\s]', '', desc).split(' ')

    # находим совпадения в запросе и описание, также делаем проверку на длину запроса
    def check_desc(self, query):
        if set(query[0].split(' ')).issubset(set(self.desc)) and len(query[0]) >= 3:
            return True

        else:
            return False

    # проверяет существование товара по опреденному запросу
    # если есть, то возвращается true и наоборот
    # метод используется в качестве аргумента для встроенной функции filter
    def check_existence(self, brand_ids):
        if self.nmid in brand_ids:
            return True
        else:
            return False

    # возвращает рекламное место товара
    def check_ad(self, ad_ids: list[str]):

        if self.nmid in ad_ids:
            return ad_ids.index(self.nmid) + 1

        else:
            return 0

    # проверяет место товара среди первых 10 страниц выдачи
    def checkFirstTenPages(self, ids):

        if self.nmid in ids:
            # индексация списков начинается с нуля
            # поэтому для получения места выдачи +1
            return ids.index(self.nmid) + 1

        else:
            return 0

    # возвращает имя топ категории
    def checkTopCategory(self, category_id, subject_base):
        if category_id in subject_base:
            return subject_base[category_id]
        else:
            return None




# класс управленец, тупа директор по продаванию говна
# менеджерит создание ссылок, получение и сравнение данных, а также их выдачу
# папа
class Indexer:

    # создаем нужные объекты классов, инициализируем nmid и seller id (нахуй потом сходит скорее всего этот сейлер id)
    # также сразу ищет совпадения среди запросов и описания товара
    def __init__(self, nmid):
        self.nmid = nmid
        self.url_operator = URLOperator()
        self.data_collector = DataCollector()
        self.__search_common()

    # ищет совпадения по описанию и всем запросам
    def __search_common(self):
        file = open(os.path.join(CURRENT_DIR, 'requests.csv'), encoding='utf-8')
        reader = csv.reader(file)

        card_url = self.url_operator.create_card_url(self.nmid)
        detail_card_url = self.url_operator.create_nmid_detail_url(self.nmid)

        card_info = self.data_collector.get_card_info(url=card_url)
        detail_info = self.data_collector.get_brand_and_name(detail_card_url)
        full_info = ' '.join([card_info, detail_info])

        self.checker = Checker(self.nmid, full_info)
        self.resulted_queries = filter(self.checker.check_desc, reader)

    # возвращает id бренда
    def __get_brand_id(self):
        detail_url = self.url_operator.create_nmid_detail_url(self.nmid)
        return self.data_collector.get_brand_id(detail_url)



    # получает глубину запроса
    def __get_req_depth(self, query):
        url = self.url_operator.create_query_req_depth_url(query=query)
        total = self.data_collector.get_req_depth(url)
        return total


    # сущетвует/не существует товар в выдаче по запросу
    def __get_existence(self, query):
        brand_id = self.__get_brand_id()
        url = self.url_operator.create_filtered_by_brand_url(query=query, brand_id=brand_id)
        brand_products = self.data_collector.get_query_by_brand(url)
        existence = self.checker.check_existence(brand_products)
        return existence

    # получает информацию по рекламе
    def __get_ad_info(self, query):
        url = self.url_operator.create_ad_url(query)
        ad_products = self.data_collector.get_ad(url)
        ad_place = self.checker.check_ad(ad_products)
        return {
            'ad_spots': len(ad_products),
            'ad_place': ad_place
        }

    # проверка первых 10 страниц выдачи
    def __get_place(self, query):
        url = self.url_operator.create_query_url_template(query=query)
        products = self.data_collector.get_query(url)
        place = self.checker.checkFirstTenPages(products)
        return place

    # получаем топ категорию по рекламе или по наибольшему количеству товаров в категории
    # или хуй - в случае сильно конченного запроса возвращается сообщение о ошибке
    def __getTopCategory(self, query):
        subject_base_url = self.url_operator.subject_base_url
        ad_info_url = self.url_operator.create_ad_url(query)
        subject_base = self.data_collector.get_subject_base(subject_base_url)
        top_category_id = self.data_collector.get_top_category(ad_info_url)
        top_category = self.checker.checkTopCategory(top_category_id, subject_base)

        if top_category:
            return top_category
        else:
            query_category_url = self.url_operator.create_query_categories_url(query)
            return self.data_collector.get_query_top_category(query_category_url)

    def iterate_total_and_top_category(self):
        for query in self.resulted_queries:
            keywords = query[0]
            frequency = query[1]
            top_category = self.__getTopCategory(keywords)
            req_depth = self.__get_req_depth(keywords)

            yield {
                'nmid': self.nmid,
                'frequency': frequency,
                'keywords': keywords,
                'top_category': top_category,
                'req_depth': req_depth,
            }
    # заключительный метод-генератор, возвращает словарь со всеми необходимыми данными
    def iterate_queries(self):
        for query in self.resulted_queries:

            keywords = query[0]
            frequency = query[1]
            top_category = self.__getTopCategory(keywords)
            req_depth = self.__get_req_depth(keywords)
            existence = self.__get_existence(keywords)

            # если товар есть в поисковой выдаче, то собираем данные по рекламе, месту и топу выдачи
            if existence:
                ad_info = self.__get_ad_info(keywords)
                ad_spots = ad_info['ad_spots']
                ad_place = ad_info['ad_place']
                place = self.__get_place(keywords)
                # __get_ad_info возвращает 0, если товар не в первой 1000 товаров
                # если товару присвоено место, то считаем процент
                if place and req_depth != 0:
                    percent = (place / req_depth) * 100
                    spot_req_depth = str(round(percent, 2)).replace('.', ';')

                # если места нет, то и не считаем ниче
                else:
                    spot_req_depth = place

            # если товара вообще нет в поисковой выдаче, то ниче не счяитаем, везде None значения
            else:
                ad_spots, ad_place, place, spot_req_depth = [None] * 4

            # возвращаем словарь со всеми значения
            yield {
                'nmid': self.nmid,
                'top_category': top_category,
                'keywords': keywords,
                'frequency': frequency,
                'req_depth': req_depth,
                'existence': existence,
                'place': place,
                'spot_req_depth': spot_req_depth,
                'ad_spots': ad_spots,
                'ad_place': ad_place
            }


class SeoCollector:
    """Парсит seo данные по :query запросу из wildberries"""

    def __init__(self, query, depth):
        self.query = query
        self.depth = depth

        self.url_operator = URLOperator()
        self.data_collector = DataCollector()

    def run(self):
        # indexeres = []

        # Если запись в базе данных уже создана, то мы не выполняем таску
        # if models.SeoReport.objects.filter(
        #    query=self.query, 
        #    depth=self.depth
        # ).first() is None:

        query_obj = models.SeoReport.objects.create(
            query=self.query, 
            depth=self.depth
        )
        url_first_ten_pages_products = URLOperator().create_query_url_template(
            query=self.query
        )
        products_id = DataCollector().get_query(url_first_ten_pages_products)

        for product_id in products_id[:self.depth]:
            indexer = Indexer(product_id)
            # Получаем количество товаров и топовую категорию по каждому
            # тексту, подходящего по критериям класса Checker
            for indexer_data in indexer.iterate_total_and_top_category():
                models.SeoReportData.objects.create(
                    **indexer_data, 
                    query=query_obj,
                )

        query_obj.completed = True
        query_obj.save()
